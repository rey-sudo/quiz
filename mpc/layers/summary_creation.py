import logging
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import ARTICLES_DIR, SUMMARY_PATH, LISTS_DIR

logger = logging.getLogger("rich")

def create_summary():
    """
    Iterates .md files from articles and lists folders in numeric order
    and generates multiple .xlsx files, one per every 100 rows.
    """
    def get_md_files(folder) -> dict:
        files = {}
        for f in Path(folder).glob("*.md"):
            match = re.search(r'(\d+)', f.stem)
            if match:
                number = match.group(1).zfill(4)
                files[number] = f
        return files

    def create_workbook_with_headers():
        wb = Workbook()
        wb.properties.defaultThemeVersion = "124226"
        wb.properties.language = "es-CO"
        ws = wb.active
        ws.title = "Summary"
        
        
        headers = ["Numero_Articulo", "Contenido_Articulo", "Lists"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 80

        return wb

    articles = get_md_files(ARTICLES_DIR)
    lists = get_md_files(LISTS_DIR)
    all_numbers = sorted(set(articles.keys()) | set(lists.keys()))

    output_dir = SUMMARY_PATH
    os.makedirs(output_dir, exist_ok=True)

    CHUNK_SIZE = 100
    output_paths = []

    for chunk_idx, chunk_start in enumerate(range(0, len(all_numbers), CHUNK_SIZE)):
        chunk_numbers = all_numbers[chunk_start:chunk_start + CHUNK_SIZE]

        # Name: summary_001-100.xlsx, summary_101-200.xlsx, etc.
        first_num = int(chunk_numbers[0])
        last_num = int(chunk_numbers[-1])
        output_path = output_dir / f"summary_{first_num:04d}-{last_num:04d}.xlsx"

        wb = create_workbook_with_headers()
        ws = wb.active

        for row_idx, number in enumerate(chunk_numbers, 2):
            article_content = articles[number].read_text(encoding="utf-8").strip() if number in articles else ""
            list_content = lists[number].read_text(encoding="utf-8").strip() if number in lists else ""

            cell_num = ws.cell(row=row_idx, column=1, value=number)
            cell_num.alignment = Alignment(horizontal="center")
            cell_num.font = Font(name="Arial", size=12)

            cell_article = ws.cell(row=row_idx, column=2, value=article_content)
            cell_article.alignment = Alignment(wrap_text=True, vertical="top")
            cell_article.font = Font(name="Arial", size=12)

            cell_list = ws.cell(row=row_idx, column=3, value=list_content)
            cell_list.alignment = Alignment(wrap_text=True, vertical="top")
            cell_list.font = Font(name="Arial", size=12)
            

        wb.save(output_path)
        output_paths.append(output_path)
        logger.info(f"Saved: {output_path.name} ({len(chunk_numbers)} rows)")

    return output_paths